# coding:utf-8
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QFileDialog
from qfluentwidgets import FluentIcon as FIF
from qfluentwidgets import qconfig, ScrollArea, PrimaryPushButton, InfoBar, InfoBarPosition, PushButton
from .common.style_sheet import StyleSheet
from .tools.warp_export import warpExport, WarpExport
import pyperclip
import json
import markdown
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class WarpInterface(ScrollArea):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.view = QWidget(self)
        self.vBoxLayout = QVBoxLayout(self.view)
        self.titleLabel = QLabel(self.tr("워프 기록"), self)

        self.updateBtn = PrimaryPushButton(FIF.SYNC, "데이터 업데이트", self)
        self.updateFullBtn = PushButton(FIF.SYNC, "전체 데이터 업데이트", self)
        self.importBtn = PushButton(FIF.PENCIL_INK, "데이터 가져오기", self)
        self.exportBtn = PushButton(FIF.SAVE_COPY, "데이터 내보내기", self)
        self.exportExcelBtn = PushButton(FIF.SAVE_COPY, "Excel로 내보내기", self)
        self.copyLinkBtn = PushButton(FIF.SHARE, "링크 복사", self)
        self.clearBtn = PushButton(FIF.DELETE, "초기화", self)
        self.warplink = None

        self.stateTooltip = None

        self.contentLabel = QLabel(parent)

        qconfig.themeChanged.connect(self.setContent)

        self.setContent()

        self.__initWidget()
        self.__connectSignalToSlot()

    def __initWidget(self):
        self.titleLabel.move(36, 30)
        self.updateBtn.move(35, 80)
        self.updateFullBtn.move(150, 80)
        self.importBtn.move(293, 80)
        self.exportBtn.move(408, 80)
        self.exportExcelBtn.move(523, 80)
        self.copyLinkBtn.move(638, 80)
        self.copyLinkBtn.setEnabled(False)
        self.clearBtn.move(753, 80)

        self.view.setObjectName('view')
        self.setViewportMargins(0, 120, 0, 20)
        self.setObjectName('warpInterface')
        self.contentLabel.setObjectName('contentLabel')
        self.titleLabel.setObjectName('warpLabel')
        StyleSheet.WARP_INTERFACE.apply(self)

        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setWidget(self.view)
        self.setWidgetResizable(True)
        self.contentLabel.setWordWrap(True)
        self.contentLabel.setMaximumWidth(800)

        # self.vBoxLayout.setSpacing(8)
        # self.vBoxLayout.setAlignment(Qt.AlignTop)
        # self.vBoxLayout.addWidget(self.titleLabel, 0, Qt.AlignTop)

        # self.vBoxLayout.setSpacing(28)
        self.vBoxLayout.setContentsMargins(36, 0, 36, 0)
        self.vBoxLayout.addWidget(self.contentLabel, 0, Qt.AlignTop)

    def __connectSignalToSlot(self):
        self.updateBtn.clicked.connect(self.__onUpdateBtnClicked)
        self.updateFullBtn.clicked.connect(self.__onUpdateFullBtnClicked)
        self.importBtn.clicked.connect(self.__onImportBtnClicked)
        self.exportBtn.clicked.connect(self.__onExportBtnClicked)
        self.exportExcelBtn.clicked.connect(self.__onExportExcelBtnClicked)
        self.copyLinkBtn.clicked.connect(self.__onCopyLinkBtnClicked)
        self.clearBtn.clicked.connect(self.__onClearBtnClicked)

    def __onUpdateBtnClicked(self):
        warpExport(self)

    def __onUpdateFullBtnClicked(self):
        warpExport(self, "full")

    def __onImportBtnClicked(self):
        try:
            path, _ = QFileDialog.getOpenFileName(self, "SRGF 데이터 형식 가져오기 지원", "", "붕괴: 스타레일 워프 기록 파일 (*.json)")
            if not path:
                return

            with open(path, 'r', encoding='utf-8') as file:
                config = json.load(file)
            warp = WarpExport(config)
            config = warp.export_data()
            with open("./warp.json", 'w', encoding='utf-8') as file:
                json.dump(config, file, ensure_ascii=False, indent=4)

            self.setContent()

            InfoBar.success(
                title=self.tr('가져오기 성공(＾∀＾●)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )
        except Exception:
            InfoBar.warning(
                title=self.tr('가져오기 실패(╥╯﹏╰╥)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onExportBtnClicked(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)
            warp = WarpExport(config)
            path, _ = QFileDialog.getSaveFileName(self, "SRGF 데이터 형식 내보내기 지원", f"SRGF_{warp.get_uid()}.json", "붕괴: 스타레일 워프 기록 파일 (*.json)")
            if not path:
                return

            with open(path, 'w', encoding='utf-8') as file:
                json.dump(config, file, ensure_ascii=False, indent=4)

            os.startfile(os.path.dirname(path))

            InfoBar.success(
                title=self.tr('내보내기 성공(＾∀＾●)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

        except Exception:
            InfoBar.warning(
                title=self.tr('내보내기 실패(╥╯﹏╰╥)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onExportExcelBtnClicked(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)
            records = config.get("list", [])
            df = pd.DataFrame(records)
            df = df[["time", "name", "item_type", "rank_type", "gacha_type"]]
            gacha_map = {
                "11": "캐릭터 이벤트 워프",
                "12": "광추 이벤트 워프",
                "21": "캐릭터 콜라보 워프",
                "22": "광추 콜라보 워프",
                "1": "상시 워프",
                "2": "초행길 워프",
            }
            df["gacha_type"] = df["gacha_type"].map(gacha_map).fillna("알 수 없음")
            df.rename(columns={
                "time": "시간",
                "name": "이름",
                "item_type": "유형",
                "rank_type": "등급",
                "gacha_type": "배너",
            }, inplace=True)
            df["총 횟수"] = range(1, len(df) + 1)
            df["스택"] = 0
            pity_counters = {}
            for idx, row in df.iterrows():
                pool = row["배너"]
                star = row["등급"]
                if pool not in pity_counters:
                    pity_counters[pool] = 0
                pity_counters[pool] += 1
                df.at[idx, "스택"] = pity_counters[pool]
                if star == "5":
                    pity_counters[pool] = 0
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel 파일로 내보내기",
                f"워프기록_{config['info'].get('uid', '알수없음')}.xlsx",
                "Excel 파일 (*.xlsx)"
            )
            if not path:
                return

            df.to_excel(path, index=False)
            wb = load_workbook(path)
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                star_cell = ws[f"D{row}"]
                try:
                    star = star_cell.value
                    if star == "5":
                        for col in ws[row]:
                            col.font = Font(color="FFA500")
                    elif star == "4":
                        for col in ws[row]:
                            col.font = Font(color="800080")
                except:
                    continue

            for column_cells in ws.columns:
                max_width = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        value = str(cell.value) if cell.value else ""
                        width = 0
                        for ch in value:
                            if u'\u4e00' <= ch <= u'\u9fff' or u'\uAC00' <= ch <= u'\uD7A3':
                                width += 2
                            else:
                                width += 1
                        if width > max_width:
                            max_width = width
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_width + 2

            wb.save(path)
            os.startfile(os.path.dirname(path))

            InfoBar.success(
                title=self.tr('내보내기 성공(＾∀＾●)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

        except Exception:
            InfoBar.warning(
                title=self.tr('내보내기 실패(╥╯﹏╰╥)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onCopyLinkBtnClicked(self):
        try:
            pyperclip.copy(self.warplink)
            InfoBar.success(
                title=self.tr('복사 성공(＾∀＾●)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )
        except Exception:
            InfoBar.warning(
                title=self.tr('복사 실패(╥╯﹏╰╥)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onClearBtnClicked(self):
        try:
            os.remove("./warp.json")
            self.setContent()
            InfoBar.success(
                title=self.tr('초기화 완료(＾∀＾●)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )
        except Exception as e:
            print(e)
            InfoBar.warning(
                title=self.tr('초기화 실패(╥╯﹏╰╥)'),
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def setContent(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)

            warp = WarpExport(config)
            if qconfig.theme.name == "DARK":
                content = warp.data_to_html("dark")
            else:
                content = warp.data_to_html("light")
            self.clearBtn.setEnabled(True)
            self.exportBtn.setEnabled(True)
            self.exportExcelBtn.setEnabled(True)
        except Exception as e:
            content = "워프 기록이 비어 있습니다. 게임 내 워프 기록을 먼저 연 다음 '데이터 업데이트'를 클릭하세요.\n\nStarRail Warp Export 또는 Starward 등 SRGF 데이터 형식을 지원하는 다른 앱에서 데이터를 가져올 수도 있습니다.\n\n'링크 복사' 기능은 미니 프로그램이나 다른 소프트웨어에서 사용할 수 있습니다."
            self.clearBtn.setEnabled(False)
            self.exportBtn.setEnabled(False)
            self.exportExcelBtn.setEnabled(False)
        self.contentLabel.setText(content)